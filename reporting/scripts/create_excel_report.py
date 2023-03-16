#!/usr/bin/env python3

import sys
import os
# insert path to be able to load the files in the lib directory
sys.path.insert(0, os.path.dirname(__file__) + "/../lib")

# for 
from os import path
# for parsing logs
import re
# for getting weekdays
import time
# for date parsing and handling
from datetime import datetime, timedelta
# for getting correct month days
import calendar
# for iterating through all possible combinations of logs
import itertools
# for proper argument handling
import argparse
# for working with xlsx files
import openpyxl
import openpyxl.styles

# for getting product info
from product import products

parser = argparse.ArgumentParser(
                    prog = 'create_excel_report.py',
                    description = 'Creates a monthly or weekly report from the DHuS software stack by reading the log files.')
parser.add_argument('timespan', type=str, help='W for weekly report, M for monthly report')


def get_logs_from_file(file_name):
    logs = []

    try: content = open(file_name, 'r').readlines()
    except:
        print("FILE:", file_name, "not found")
        return logs

    for line in content:
        timestamp = None

        if "download by user" in line and "completed in" in line:
            # example log line for Frontend
            # [3.0.8-osf][2022-12-21 00:01:30,396][INFO ] Product '191b2235-9324-4275-bbc1-addfd0cb5dce' (S2B_MSIL2A_20221218T210809_N0509_R014_T03LZH_20221218T222128) download by user 'dhr_uk_stfc' completed in 244ms -> 9485208 (DownloadActionRecordListener.java:65 - http-nio-8092-exec-8)
            timestamp = datetime.strptime(re.search(r"[\d]{4}-[\d]{2}-[\d]{2} [\d]{2}:[\d]{2}:[\d]{2},[\d]{3}", line).group(0), "%Y-%m-%d %H:%M:%S,%f")
            uuid = re.search(r"'[0-9a-z-]+'", line).group(0)[1:-1]
            product_name = re.search(r"\([0-9A-Z_]+\)", line).group(0)[1:-1]
            download_time = int(re.search(r"[\d]+ms", line).group(0)[:-2])
            size = int(re.search(r"-> [\d]+", line).group(0)[3:])
            user_name = re.search(r"user '.*'", line).group(0)[6:-1]
            remote_site = ""

        if "successfully downloaded" in line:
            # example log line for Backend
            # [3.0.8-osf][2022-12-20 02:52:23,774][INFO ] Product 'S1A_IW_RAW__0SDV_20221219T224545_20221219T224617_046407_058F35_B2D2.zip' (1565309617 bytes compressed) successfully downloaded from https://sentinelhub2.met.no/odata/v1/Products('53f2aa9c-4104-43e3-b399-4678e24fae23')/$value in 69133 ms (DownloadableProduct.java:341 - Product Download)

            timestamp = datetime.strptime(re.search(r"[\d]{4}-[\d]{2}-[\d]{2} [\d]{2}:[\d]{2}:[\d]{2},[\d]{3}", line).group(0), "%Y-%m-%d %H:%M:%S,%f")
            uuid = re.search(r"'[0-9a-z-]+'", line).group(0)[1:-1]
            product_name = re.search(r"'[0-9A-Z_]+", line).group(0)[1:]
            download_time = int(re.search(r"[\d]+ ms", line).group(0)[:-3])
            size = int(re.search(r"[\d]+ bytes", line).group(0)[:-6])
            remote_site = re.search(r"from https://[a-z\.0-9-]+", line).group(0)[13:]
            user_name = ""

        if timestamp:
            logs.append({
                'timestamp': timestamp,
                'uuid': uuid,
                'product_name': product_name,
                'download_time': download_time,
                'size': size,
                'remote_site': remote_site,
                'user_name': user_name,
            })

    return logs

def get_logs():
    logs = []
    for dhus_directory in DHUS_DIRECTORIES:
        log_directory = LOG_DIRECTORY.format(dhus_directory=dhus_directory)

        if path.exists(log_directory):
            if ARG.timespan == 'M':
                for day in range(1, calendar.monthrange(YEAR, MONTH)[1] + 1):
                    log_file = '{log_directory}/dhus-{year}-{month:02.0f}-{day:02.0f}.log'.format(log_directory=log_directory, year=YEAR, month=MONTH, day=day)
                    logs.extend(get_logs_from_file(log_file))
            
            if ARG.timespan == 'W':
                # get list of weekdays
                startdate = time.asctime(time.strptime('%d %d 1' % (YEAR, WEEK), '%Y %W %w'))
                startdate = datetime.strptime(startdate, '%a %b %d %H:%M:%S %Y')
                dates = [startdate.strftime('%Y-%m-%d')]
                for i in range(1, 7):
                    day = startdate + timedelta(days=i)
                    dates.append(day.strftime('%Y-%m-%d'))

                # iterate through weekdays one by one
                for d in dates:
                    log_file = '{log_directory}/dhus-{d}.log'.format(log_directory=log_directory, d=d)
                    logs.extend(get_logs_from_file(log_file))

        else: print("PATH:", log_directory, "does not exist")
    
    return logs

def write_report(logs):
    # put all product groups into one list for iteration
    product_names = []
    for product in products.values():
        product_names.extend(product["names"])
    
    # get distinct list of product names
    product_names = set(product_names)
    # get distinct list of remote sites
    remote_sites = set([log['remote_site'] for log in logs])
    # get distinct list of user names
    user_names = set([log['user_name'] for log in logs])

    retrieved_stats = []
    distributed_stats = []
    bandwidth_stats = []
    # cross product is the easiest and fastest way to comb through all combinations
    # and get the relevant data grouped together
    # - product_names are needed in all stats cases (retrieved, distributed, bandwidth)
    # - remote_sites are needed for the retrieved stats
    # - user_names are needed for distributed stats
    # see also the get_logs_from_file(file_name) function above as to when which variable is set
    for p in itertools.product(product_names, remote_sites, user_names):
        product_name = p[0]
        remote_site = p[1]
        user_name = p[2]

        # get all logs relevant to this product_group AND remote_site AND user_name
        product_logs = [log for log in logs
                            if all([True if p in log['product_name']
                                         else False for p in product_name.split("..")])
                               and
                               log['remote_site'] == remote_site
                               and
                               log['user_name'] == user_name]

        # get number of unique products
        nb_products = len(set([p['product_name'] for p in product_logs]))

        # get size in TB, download_time in sec and bandwidth in Mbps
        size = sum([p['size'] for p in product_logs])
        download_time = sum([p['download_time'] for p in product_logs]) / 1000
        bandwidth = size / download_time if download_time > 0 else 0

        # if remote_site is not empty => retrieved_stats
        # else => distributed_stats
        reporting_period = ""
        if ARG.timespan == 'M':
            reporting_period = "{year}-{month:02.0f}-01 to {year}-{month:02.0f}-{day:02.0f}".format(year=YEAR, month=MONTH, day=calendar.monthrange(YEAR, MONTH)[1])
        if ARG.timespan == 'W':
            # get list of weekdays
            start_date = time.asctime(time.strptime('%d %d 1' % (YEAR, WEEK), '%Y %W %w'))
            start_date = datetime.strptime(start_date, '%a %b %d %H:%M:%S %Y')
            end_date = (start_date + timedelta(days=6)).strftime('%Y-%m-%d')
            start_date = start_date.strftime('%Y-%m-%d')

            reporting_period = "{start_date} to {end_date}".format(start_date=start_date, end_date=end_date)
        if remote_site:
            # only put in excel sheet if number of products != 0
            if nb_products != 0:
                retrieved_stats.append({
                    'key': remote_site + "-" + product_name,
                    'from': "https://" + remote_site + "/odata/v1",
                    'to': "ZAMG",
                    'product_type': product_name,
                    'nb_products': nb_products,
                    'size': round(size / (2**40), 3), # in TB 
                    'reporting_period': reporting_period
                })
        else:
            # only put in excel sheet if number of products != 0
            if nb_products != 0:
                distributed_stats.append({
                    'key': "'" + user_name + "'-" + product_name,
                    'from': "ZAMG",
                    'to': user_name,
                    'product_type': product_name,
                    'nb_products': nb_products,
                    'size': round(size / (2**40), 3), # in TB
                    'reporting_period': reporting_period
                })
            # only put in sheet if bandwidth != 0
            if bandwidth != 0:
                bandwidth_stats.append({
                    'key': "'" + user_name + "'-" + product_name,
                    'from': "ZAMG",
                    'to': user_name,
                    'product_type': product_name,
                    'bandwidth': round(bandwidth / (2**20) * 8, 2), # in Mbps
                    'reporting_period': reporting_period
                })

    # get xlsx workbook and necessary worksheets
    cur_dir = path.dirname(__file__)
    workbook = openpyxl.load_workbook(cur_dir + '/../reports/report.xlsx')
    worksheet_retrieved = workbook['RetrieveStats']
    worksheet_distributed = workbook['DistributedStats']
    worksheet_bandwidth = workbook['BandwidthStats']

    # set correct headers for all three sheets
    header_font = openpyxl.styles.Font(bold=True)
    headers = [ "key", "from", "to", "product type", "nb products", "size", "reporting period" ]
    headers_bandwidth = [ "key", "from", "to", "product type", "bandwidth (Mbps)", "reporting period", "" ]
    for i in range(len(headers)):
        worksheet_retrieved.cell(row=1, column=i+1).font = header_font
        worksheet_retrieved.cell(row=1, column=i+1).value = headers[i]
        worksheet_distributed.cell(row=1, column=i+1).font = header_font
        worksheet_distributed.cell(row=1, column=i+1).value = headers[i]
        worksheet_bandwidth.cell(row=1, column=i+1).font = header_font
        worksheet_bandwidth.cell(row=1, column=i+1).value = headers_bandwidth[i]

    # sort retrieved_stats list and
    # fill stats into worksheet
    retrieved_stats = sorted(retrieved_stats, key=lambda x: x['key'])
    for stat in retrieved_stats:
        worksheet_retrieved.append([stat['key'], stat['from'], stat['to'], stat['product_type'], stat['nb_products'], stat['size'], stat['reporting_period']])

    # sort distributed_stats list and
    # fill stats into worksheet
    distributed_stats = sorted(distributed_stats, key=lambda x: x['key'])
    for stat in distributed_stats:
        worksheet_distributed.append([stat['key'], stat['from'], stat['to'], stat['product_type'], stat['nb_products'], stat['size'], stat['reporting_period']])

    # sort bandwidth_stats list and
    # fill stats into worksheet
    bandwidth_stats = sorted(bandwidth_stats, key=lambda x: x['key'])
    for stat in bandwidth_stats:
        worksheet_bandwidth.append([stat['key'], stat['from'], stat['to'], stat['product_type'], stat['bandwidth'], stat['reporting_period']])

    if ARG.timespan == 'M':
        workbook.save(cur_dir + '/../reports/excel_reports/{year}-m{month:02.0f}_report.xlsx'.format(year=YEAR, month=MONTH))

    if ARG.timespan == 'W':
        workbook.save(cur_dir + '/../reports/excel_reports/{year}-w{week:02.0f}_report.xlsx'.format(year=YEAR, week=WEEK))

# get argument
ARG = parser.parse_args()

# adjust these variables accordingly to your setup
DHUS_DIRECTORIES = ["dhr-logs", "s1-logs", "s2a-l1c-logs", "s2b-l1c-logs", "s2-l2a-logs", "s3-logs", "s5-logs"]
LOG_DIRECTORY = "YOUR_LOGS_PATH/{dhus_directory}"

# get last weeks report
# therefore set WEEK and YEAR variables accordingly
if ARG.timespan == 'W':
    WEEK = datetime.now().isocalendar()[1] - 1 if datetime.now().isocalendar()[1] != 1 else 52
    YEAR = datetime.now().year if WEEK != 52 else datetime.now().year - 1
# get last months report
# therefore set MONTH and YEAR variables accordingly
if ARG.timespan == 'M':
    MONTH = datetime.now().month - 1 if datetime.now().month != 1 else 12
    YEAR = datetime.now().year if MONTH != 12 else datetime.now().year - 1

logs = get_logs()
write_report(logs)